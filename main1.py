
from openpyxl import Workbook, load_workbook
import os
import datetime

data = []
school = []
personal = []

def convert_to_delta(s):
    try:
        minutes, seconds = s.split(':')
        seconds, hundredths = seconds.split('.')
        return datetime.timedelta(
            minutes=int(minutes),
            seconds=int(seconds),
            microseconds=int(hundredths) * 10_000
        )
    except Exception as e:
        print(f"Ошибка преобразования времени: {s}")
        raise e

def convert_to_string(t):
    total_seconds = t.seconds + t.days * 86400
    minutes, seconds = divmod(total_seconds, 60)
    hundredths = int(t.microseconds / 10_000)
    return f"{minutes:02}:{seconds:02}.{hundredths:02}"

def check_file():
    if not os.path.isfile("input.xlsx"):
        wb = Workbook()
        wb.remove(wb["Sheet"])
        for i in range(5):
            sheet = wb.create_sheet(f"Забег{i + 1}")
            for _ in range(5):
                sheet.append(["Имя", "Школа", "00:00.00"])
        wb.save("input.xlsx")
        wb.close()
        print("Создан шаблон input.xlsx. Заполни его и перезапусти скрипт.")
        exit()

def load_data():
    wb = load_workbook("input.xlsx")
    for sheet in wb.sheetnames:
        for row in wb[sheet].iter_rows(values_only=True):
            if all(row[:3]):
                try:
                    time = convert_to_delta(row[2])
                    data.append([row[0], row[1], time])
                except:
                    continue
    wb.close()

def sort_data():
    sh_dict = {}
    for name, school_name, time in data:
        if school_name not in sh_dict:
            sh_dict[school_name] = []
        sh_dict[school_name].append(time)

    for school_name in sh_dict:
        sh_dict[school_name] = sorted(sh_dict[school_name])[:8]

    global school, personal
    school = sorted(
        [(k, sum(v, datetime.timedelta())) for k, v in sh_dict.items()],
        key=lambda x: x[1]
    )

    personal = sorted(data, key=lambda x: x[2])
    personal = [(x[0], x[1], convert_to_string(x[2])) for x in personal]

def save_reports():
    wb = Workbook()
    wb.remove(wb["Sheet"])

    ws1 = wb.create_sheet("Школы")
    for row in school:
        ws1.append([row[0], convert_to_string(row[1])])

    ws2 = wb.create_sheet("Персональный")
    for row in personal:
        ws2.append(row)

    wb.save("output.xlsx")
    wb.close()

if __name__ == "__main__":
    check_file()
    load_data()
    sort_data()
    save_reports()
    print("Готово! Проверь output.xlsx.")
