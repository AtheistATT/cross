from openpyxl import *
import os
import pdb
import datetime

data = []
school = []
personal = []


def convert_to_delta(s):
    minutes, seconds = s.split(':')
    seconds, hundredths = seconds.split('.')
    return datetime.timedelta(
        minutes=int(minutes),
        seconds=int(seconds),
        microseconds=int(hundredths) * 10_000  # 0.01 сек = 10_000 микросекунд
    )

def convert_to_string(t):
    total_seconds = t.seconds + t.days * 86400
    minutes, seconds = divmod(total_seconds, 60)
    hundredths = int(t.microseconds / 10_000)
    return f"{minutes:02}:{seconds:02}.{hundredths:02}"

def check_file():
    if not os.path.isfile("input.xlsx"):
        wb = Workbook()
        wb.remove(wb["Sheet"])

        for i in range(5):
            wb.create_sheet(f"Забег{i + 1}")
            for _ in range(5):
                wb[(f"Забег{i + 1}")].append(["Имя","Школа","00:00.00"])
        wb.save("input.xlsx")
        wb.close()
        exit()

def load_data():
    wb = load_workbook("input.xlsx")

    for sheet in wb.sheetnames:
        s = wb[sheet]
        for row in s.iter_rows(values_only=True):
            data.append(list(row))

    for x in data:
        x[2] = convert_to_delta(x[2]) 

def sort_data():

    sh_dict = {}

    data8 = data

    for s in data8:
        if s[1] not in sh_dict:
            sh_dict[s[1]] = []
        sh_dict[s[1]] += [s[2]]


    for x in sh_dict.keys():
        sh_dict[x] = sorted(sh_dict[x])
        sh_dict[x] = sh_dict[x][:8]
    

    global school
    global personal
    
    school = sorted([(k, sum(v, datetime.timedelta())) for k, v in sh_dict.items()], key=lambda x: x[1])

    personal = sorted(data, key=lambda x: x[2])
    personal = [(x[0], x[1], convert_to_string(x[2])) for x in personal]

    
def save_reports():
    wbo = Workbook()

    wbo.remove(wbo["Sheet"])
    wbo.create_sheet("Школы")

    for row in school:
        wbo["Школы"].append(row)


    wbo.create_sheet("Персональный")

    for row in personal:
        wbo["Персональный"].append(row)

    wbo.save("output.xlsx")
    wbo.close()


check_file()
load_data()
sort_data()
save_reports()
